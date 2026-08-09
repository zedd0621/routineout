import io
import os
import re
import math
import calendar
from datetime import datetime, timedelta

try:
    from openpyxl import load_workbook
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas as pdfcanvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

# ── 급여 상수 (원본 flask_app.py 값 그대로) ──
WAGE_BASE = 12100
WAGE_MAIN_BONUS = 17900
WAGE_SUB_BONUS = 7900
EARLY_LATE_BONUS = 6050
HEADCOUNT_BONUS_HOURLY = 2000
HEADCOUNT_MIN = 20
SATURDAY_MULT = 1.5
NIGHT_MULT = 1.5
NIGHT_START_HOUR = 18

WEEKLY_HOLIDAY_MIN_HOURS = 15
WEEKLY_HOLIDAY_CAP_HOURS = 40

INSURANCE_HOURS_MIN = 60
EMPLOYMENT_INSURANCE_EXEMPT_AGE = 65

ASSETS_FOLDER = os.path.join(os.path.dirname(__file__), 'assets')
FONT_PATH = os.path.join(ASSETS_FOLDER, 'NanumGothic.ttf')
FONT_NAME = 'NanumGothic'

PAYSLIP_FONT_OK = False
if REPORTLAB_OK and os.path.exists(FONT_PATH):
    try:
        pdfmetrics.registerFont(TTFont(FONT_NAME, FONT_PATH))
        PAYSLIP_FONT_OK = True
    except Exception:
        PAYSLIP_FONT_OK = False


def ceil_15min(td_seconds):
    minutes = td_seconds / 60
    if minutes <= 0:
        return 0
    return math.ceil(minutes / 15) * 15


def load_instructor_list(paycheck_folder):
    """강사명단 엑셀 로딩 -> {이름: {'birth': 'YYYY-MM-DD', 'tel': '010-xxxx-xxxx'}}"""
    path = os.path.join(paycheck_folder, 'instructor_list.xlsx')
    if not os.path.exists(path) or not OPENPYXL_OK:
        return {}
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb[wb.sheetnames[0]]
        result = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[1]
            tel = row[2]
            full = row[0]
            if not name:
                continue
            birth = ''
            if full:
                m = re.search(r'(\d{4}-\d{2}-\d{2})', str(full))
                if m:
                    birth = m.group(1)
            result[str(name).strip()] = {
                'birth': birth,
                'tel': str(tel).strip() if tel else ''
            }
        return result
    except Exception:
        return {}


def load_worklog(paycheck_folder, target_month):
    """시프티 근무일지 엑셀 로딩. worklog_YYYY-MM.xlsx"""
    path = os.path.join(paycheck_folder, f'worklog_{target_month}.xlsx')
    if not os.path.exists(path) or not OPENPYXL_OK:
        return {}
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        result = {}

        has_headcount_cols = False
        try:
            header = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
            h24 = str(header[24]) if len(header) > 24 and header[24] else ''
            h25 = str(header[25]) if len(header) > 25 and header[25] else ''
            has_headcount_cols = ('교육인원' in h24) and ('교육인원' in h25)
        except Exception:
            has_headcount_cols = False

        for row in ws.iter_rows(min_row=5, values_only=True):
            name = row[1]
            branch = row[2]
            role = row[3]
            date_val = row[4]
            sched_start = row[8]
            sched_end = row[9]
            actual_in = row[10]
            actual_out = row[12]
            sched_hours_td = row[20]
            total_hours_td = row[22]
            lecture_hours_td = row[23]

            if not name or date_val is None:
                continue

            def to_hours(v):
                if v is None:
                    return 0.0
                if hasattr(v, 'total_seconds'):
                    return v.total_seconds() / 3600
                if hasattr(v, 'hour'):
                    return v.hour + v.minute / 60 + v.second / 3600
                try:
                    return float(v) * 24
                except Exception:
                    return 0.0

            sched_hours = to_hours(sched_hours_td)
            lecture_hours = to_hours(lecture_hours_td)

            role_str = str(role).strip() if role else ''
            if '주강사' in role_str:
                bonus_rate = WAGE_MAIN_BONUS
            elif '보조강사' in role_str or '보조' in role_str:
                bonus_rate = WAGE_SUB_BONUS
            else:
                bonus_rate = 0

            is_saturday = date_val.weekday() == 5 if hasattr(date_val, 'weekday') else False

            night_hours = 0.0
            try:
                if hasattr(sched_start, 'hour') and hasattr(sched_end, 'hour'):
                    night_from = sched_end.replace(hour=NIGHT_START_HOUR, minute=0,
                                                   second=0, microsecond=0)
                    if sched_start > night_from:
                        night_from = sched_start
                    if sched_end > night_from:
                        night_hours = (sched_end - night_from).total_seconds() / 3600
            except Exception:
                night_hours = 0.0
            night_hours = min(max(night_hours, 0.0), sched_hours)

            if is_saturday:
                night_hours = 0.0
                base_pay = sched_hours * WAGE_BASE * SATURDAY_MULT
                lecture_pay = lecture_hours * bonus_rate * SATURDAY_MULT
            else:
                normal_hours = max(sched_hours - night_hours, 0.0)
                base_pay = (normal_hours * WAGE_BASE
                            + night_hours * WAGE_BASE * NIGHT_MULT)
                lecture_pay = lecture_hours * bonus_rate

            def _to_int(v):
                try:
                    return int(float(v))
                except (TypeError, ValueError):
                    return 0

            headcount = 0
            headcount_bonus = 0
            if has_headcount_cols:
                hc_main = _to_int(row[24]) if len(row) > 24 else 0
                hc_sub = _to_int(row[25]) if len(row) > 25 else 0
                if '주강사' in role_str:
                    headcount = hc_main
                elif '보조' in role_str:
                    headcount = hc_sub
                else:
                    headcount = max(hc_main, hc_sub)
                if headcount >= HEADCOUNT_MIN:
                    headcount_bonus = round(sched_hours * HEADCOUNT_BONUS_HOURLY)

            early_bonus = 0
            late_bonus = 0
            try:
                if actual_in is not None and sched_start is not None and hasattr(actual_in, 'timestamp'):
                    diff_sec = (sched_start - actual_in).total_seconds()
                    if diff_sec > 0:
                        rounded_min = ceil_15min(diff_sec)
                        if rounded_min >= 30:
                            early_bonus = round(EARLY_LATE_BONUS * (SATURDAY_MULT if is_saturday else 1))
                if actual_out is not None and sched_end is not None and hasattr(actual_out, 'timestamp'):
                    diff_sec = (actual_out - sched_end).total_seconds()
                    if diff_sec > 0:
                        rounded_min = ceil_15min(diff_sec)
                        if rounded_min >= 30:
                            late_bonus = round(EARLY_LATE_BONUS * (SATURDAY_MULT if is_saturday else 1))
            except Exception:
                pass

            daily_pay = round(base_pay + lecture_pay + early_bonus + late_bonus + headcount_bonus)

            date_str = date_val.strftime('%Y-%m-%d') if hasattr(date_val, 'strftime') else str(date_val)
            weekday_kr = ['월', '화', '수', '목', '금', '토', '일']
            weekday = weekday_kr[date_val.weekday()] if hasattr(date_val, 'weekday') else ''

            entry = {
                'date': date_str,
                'weekday': weekday,
                'role': role_str,
                'sched_start': sched_start.strftime('%H:%M') if hasattr(sched_start, 'strftime') else '',
                'sched_end': sched_end.strftime('%H:%M') if hasattr(sched_end, 'strftime') else '',
                'actual_in': actual_in.strftime('%H:%M') if hasattr(actual_in, 'strftime') else '',
                'actual_out': actual_out.strftime('%H:%M') if hasattr(actual_out, 'strftime') else '',
                'total_hours': round(sched_hours, 2),
                'lecture_hours': round(lecture_hours, 2),
                'base_pay': round(base_pay),
                'lecture_pay': round(lecture_pay),
                'early_bonus': early_bonus,
                'late_bonus': late_bonus,
                'headcount': headcount,
                'headcount_bonus': headcount_bonus,
                'is_saturday': is_saturday,
                'night_hours': round(night_hours, 2),
                'sched_hours': round(sched_hours, 2),
                'branch': str(branch).strip() if branch else '',
                'daily_pay': daily_pay
            }

            name_key = str(name).strip()
            result.setdefault(name_key, []).append(entry)

        for k in result:
            result[k].sort(key=lambda x: x['date'])

        return result
    except Exception:
        return {}


def prev_month_str(month_str):
    y, m = map(int, month_str.split('-'))
    if m == 1:
        return f'{y - 1}-12'
    return f'{y}-{m - 1:02d}'


def calc_weekly_holiday(entries, target_month, prior_entries=None):
    """주휴수당 계산. 주 소정근로 15h 이상이면 min(주간시간,40)/40*8*WAGE_BASE.
    월경계 주는 전달에서 이미 지급됐는지 prior_entries로 확인해 중복지급을 막는다."""

    def hours_by_week(ent_list):
        weeks = {}
        for e in (ent_list or []):
            try:
                d = datetime.strptime(e['date'], '%Y-%m-%d').date()
            except Exception:
                continue
            monday = d - timedelta(days=d.weekday())
            h = float(e.get('sched_hours', 0.0) or 0.0)
            weeks[monday] = weeks.get(monday, 0.0) + h
        return weeks

    weeks_this = hours_by_week(entries)
    weeks_prior = hours_by_week(prior_entries)

    rows = []
    total = 0
    for monday in sorted(weeks_this):
        hours = weeks_this[monday]
        qualified = hours >= WEEKLY_HOLIDAY_MIN_HOURS
        dup_skipped = False

        if monday.strftime('%Y-%m') != target_month:
            prior_hours = weeks_prior.get(monday, 0.0)
            if prior_hours >= WEEKLY_HOLIDAY_MIN_HOURS:
                dup_skipped = True

        pay = 0
        if qualified and not dup_skipped:
            pay = round(min(hours, WEEKLY_HOLIDAY_CAP_HOURS) / WEEKLY_HOLIDAY_CAP_HOURS
                        * 8 * WAGE_BASE)
        total += pay
        rows.append({
            'week_start': monday.strftime('%m-%d'),
            'week_end': (monday + timedelta(days=6)).strftime('%m-%d'),
            'hours': round(hours, 1),
            'qualified': qualified,
            'pay': pay,
            'dup_skipped': dup_skipped,
        })
    return rows, total


def judge_insurance(entries, target_month):
    """월 소정근로 60h 이상이면 4대보험, 미만이면 2대보험(고용·산재)."""
    month_entries = [e for e in entries if str(e.get('date', ''))[:7] == target_month]
    month_hours = sum(float(e.get('sched_hours', 0.0) or 0.0) for e in month_entries)
    is_four = month_hours >= INSURANCE_HOURS_MIN
    reasons = []
    if is_four:
        reasons.append('월 소정근로 {:.1f}h (기준 {}h 이상)'.format(month_hours, INSURANCE_HOURS_MIN))
    return is_four, reasons, round(month_hours, 1)


def build_payroll_summary(paycheck_folder, target_months):
    """지정한 월들의 강사별 급여를 누적 집계한다."""
    months_used = [
        m for m in target_months
        if os.path.exists(os.path.join(paycheck_folder, f'worklog_{m}.xlsx'))
    ]

    per_name_entries = {}
    for m in months_used:
        for name, entries in load_worklog(paycheck_folder, m).items():
            per_name_entries.setdefault(name, []).extend(entries)

    rows = []
    totals = {'base': 0, 'lecture': 0, 'headcount': 0, 'weekly': 0, 'total': 0}
    for name, entries in per_name_entries.items():
        base = sum(int(e.get('base_pay', 0)) + int(e.get('early_bonus', 0))
                   + int(e.get('late_bonus', 0)) for e in entries)
        lecture = sum(int(e.get('lecture_pay', 0)) for e in entries)
        headcount = sum(int(e.get('headcount_bonus', 0)) for e in entries)
        weekly = 0
        for m in months_used:
            _, wt = calc_weekly_holiday(entries, m, prior_entries=entries)
            weekly += wt
        total = base + lecture + headcount + weekly

        worked_months = sorted({str(e.get('date', ''))[:7] for e in entries if e.get('date')})
        ins_month = worked_months[-1] if worked_months else (months_used[-1] if months_used else None)
        if ins_month:
            is_four, ins_reasons, ins_hours = judge_insurance(entries, ins_month)
        else:
            is_four, ins_reasons, ins_hours = False, [], 0
        insurance_type = '4대보험' if is_four else '2대보험'

        rows.append({
            'name': name,
            'base': base,
            'lecture': lecture,
            'headcount': headcount,
            'weekly': weekly,
            'total': total,
            'insurance_type': insurance_type,
            'insurance_month': ins_month,
            'insurance_reasons': ins_reasons,
        })
        totals['base'] += base
        totals['lecture'] += lecture
        totals['headcount'] += headcount
        totals['weekly'] += weekly
        totals['total'] += total

    rows.sort(key=lambda r: r['total'], reverse=True)
    return rows, totals, months_used


def get_available_months(paycheck_folder):
    if not os.path.exists(paycheck_folder):
        return []
    months = []
    for fname in os.listdir(paycheck_folder):
        m = re.match(r'worklog_(\d{4}-\d{2})\.xlsx', fname)
        if m:
            months.append(m.group(1))
    return sorted(months, reverse=True)


def normalize_birth(s):
    return re.sub(r'\D', '', s or '')


def is_employment_insurance_exempt_by_age(birth_str, target_month):
    """만 65세 이상이면 고용보험(실업급여분 0.9%) 공제 면제 (고용보험법 제10조 제1항)."""
    birth_digits = normalize_birth(birth_str)
    if len(birth_digits) != 8:
        return False
    try:
        by, bm, bd = int(birth_digits[:4]), int(birth_digits[4:6]), int(birth_digits[6:8])
        birth_date = datetime(by, bm, bd).date()
    except Exception:
        return False

    ty, tm = map(int, target_month.split('-'))
    last_day = calendar.monthrange(ty, tm)[1]
    ref_date = datetime(ty, tm, last_day).date()

    age = ref_date.year - birth_date.year - (
        (ref_date.month, ref_date.day) < (birth_date.month, birth_date.day)
    )
    return age >= EMPLOYMENT_INSURANCE_EXEMPT_AGE


def generate_payslip_pdf(company_name, name, target_month, base_total, lecture_total,
                          weekly_total, total_hours, lecture_hours,
                          bank_name, account_number, insurance_deduction):
    """급여 지급명세서 PDF 생성. 데모용 가상 데이터이므로 직인은 넣지 않는다."""
    buf = io.BytesIO()
    c = pdfcanvas.Canvas(buf, pagesize=A4)
    W, H = A4
    F = FONT_NAME if PAYSLIP_FONT_OK else 'Helvetica'

    total_pay = base_total + lecture_total + weekly_total
    deduction_total = insurance_deduction
    net_pay = total_pay - deduction_total

    issue_date = datetime.now().strftime('%Y년 %m월 %d일')
    issue_date_dot = datetime.now().strftime('%Y.%m.%d')

    y = H - 60
    c.setFont(F, 18)
    c.drawString(60, y, f'{issue_date_dot}  급여 지급명세서')
    c.setFont(F, 9)
    c.drawRightString(W - 60, y + 4, company_name)
    c.drawRightString(W - 60, y - 10, name)

    y -= 45
    c.setFont(F, 9)
    c.drawString(60, y, '지급')
    c.drawString(140, y, issue_date)
    if bank_name and account_number:
        c.drawRightString(W - 60, y, bank_name)
        c.drawRightString(W - 60, y - 14, account_number)

    y -= 35
    c.setFont(F, 10)
    c.drawString(60, y, '통상시급')
    c.drawString(140, y, f'{WAGE_BASE:,}원')
    c.drawString(320, y, '실수령액')
    c.setFont(F, 12)
    c.drawRightString(W - 60, y, f'{net_pay:,}')

    y -= 25
    c.line(55, y, W - 55, y)
    y -= 20
    c.setFont(F, 11)
    c.drawString(60, y, '지급합계')
    c.drawRightString(300, y, f'{total_pay:,}')
    c.drawString(320, y, '공제합계')
    c.drawRightString(W - 60, y, f'{deduction_total:,}')

    rows_left = [
        ('기본급', '시급*근무시간', base_total),
        ('강사수당', '강사수당*강의시간', lecture_total),
        ('주휴수당', '시급*주휴시간', weekly_total),
    ]
    rows_right = [
        ('국민연금', '국민연금 4.75%', 0),
        ('건강보험', '건강보험', 0),
        ('고용보험', '고용보험 0.9%', insurance_deduction),
        ('소득세', '', 0),
        ('지방소득세', '', 0),
    ]
    y -= 18
    c.setFont(F, 9)
    yy = y
    for label, formula, val in rows_left:
        c.drawString(60, yy, label)
        c.setFillGray(0.55)
        c.drawString(140, yy, formula)
        c.setFillGray(0)
        c.drawRightString(300, yy, f'{val:,}')
        yy -= 16
    yy = y
    for label, formula, val in rows_right:
        c.drawString(320, yy, label)
        c.setFillGray(0.55)
        c.drawString(400, yy, formula)
        c.setFillGray(0)
        c.drawRightString(W - 60, yy, f'{val:,}')
        yy -= 16

    y -= 120
    c.setFont(F, 9)
    c.drawString(60, y, '근무시간')
    c.drawRightString(300, y, f'{total_hours:g}')
    y -= 16
    c.drawString(60, y, '강의시간')
    c.drawRightString(300, y, f'{lecture_hours:g}')

    y = 140
    c.setFont(F, 13)
    c.drawString(60, y, company_name)
    c.setFont(F, 9)
    c.setFillGray(0.5)
    c.drawString(60, y - 16, '(데모용 가상 데이터 — 직인 생략)')
    c.setFillGray(0)
    c.setFont(F, 11)
    c.drawString(W - 220, y + 20, issue_date)

    c.showPage()
    c.save()
    buf.seek(0)
    return buf
