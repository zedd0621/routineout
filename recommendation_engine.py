import json
import os
import re

try:
    from openpyxl import load_workbook
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

from tenant_db import get_db, tenant_dir

REGIONS = ['김포', '고양', '파주', '양주', '의정부', '동두천', '연천', '포천',
           '가평', '남양주', '구리', '하남', '광주', '양평', '이천', '여주']

ALL_DAYS = ['월', '화', '수', '목', '금', '토']
ALL_TIMES = ['1교시(10:00~12:00)', '2교시(13:00~15:00)', '3교시(16:00~18:00)']
ALL_MONTHS = ['7월', '8월', '9월', '10월', '11월', '12월']


def interview_excel_path(tenant):
    return os.path.join(tenant_dir(tenant), 'interview_results.xlsx')


def load_bonus_data(tenant):
    """면접결과 관리표에서 합격자의 우대사항과 최종점수를 읽는다."""
    path = interview_excel_path(tenant)
    if not OPENPYXL_OK or not os.path.exists(path):
        return {}
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb['면접결과 관리표']
        bonus_map = {}
        for row in ws.iter_rows(min_row=7, max_row=300, values_only=True):
            name = row[17]
            bonus = row[14]
            result = row[15]
            u_val = row[20]
            if not name or result != '합격':
                continue
            bonus_str = str(bonus) if bonus else ''
            try:
                interview_score = float(u_val) if u_val is not None else 0.0
            except (TypeError, ValueError):
                interview_score = 0.0
            bonus_map[str(name).strip()] = {
                '청년': '청년' in bonus_str,
                '신규': '신규' in bonus_str,
                '취약': '취약' in bonus_str,
                'interview_score': interview_score,
            }
        return bonus_map
    except Exception:
        return {}


def calc_bonus_score(bonus_dict):
    items = []
    if bonus_dict.get('청년'): items.append(('청년', 4))
    if bonus_dict.get('취약'): items.append(('취약', 2))
    if bonus_dict.get('신규'): items.append(('신규', 2))

    if not items:
        return 0, []

    items.sort(key=lambda x: x[1], reverse=True)

    if len(items) == 1:
        total = items[0][1]
        labels = [f'{items[0][0]} +{items[0][1]}점 가산점 반영']
    else:
        top = items[0]
        rest_score = sum(v for _, v in items[1:])
        half_score = rest_score * 0.5
        total = top[1] + half_score
        rest_labels = '+'.join(f'{n}' for n, _ in items[1:])
        labels = [
            f'{top[0]} +{top[1]}점 가산점 반영',
            f'{rest_labels} +{half_score}점 가산점 반영 (중복 50% 적용)',
        ]

    return round(total, 2), labels


def parse_list(text):
    if not text:
        return []
    return [t.strip() for t in text.split(',') if t.strip()]


def parse_time_slots(times_list):
    slots = set()
    for t in times_list:
        if '1교시' in t:
            slots.add('1교시')
        if '2교시' in t:
            slots.add('2교시')
        if '3교시' in t:
            slots.add('3교시')
    return slots


def build_recommendation(tenant, center_region, center_org):
    """센터 요구사항(요일/교시)과 강사 지원 데이터를 비교하여 추천 우선순위 리스트를 반환.
    배점: 요일충족률 50점 + 교시충족률 50점 + 지망순위 최대 3점 + 면접우대사항 가산점"""
    bonus_data = load_bonus_data(tenant)

    conn = get_db(tenant)
    center_row = conn.execute(
        "SELECT * FROM responses WHERE region=? AND org=? AND confirmed=1 LIMIT 1",
        (center_region, center_org)
    ).fetchone()

    if not center_row:
        conn.close()
        return None, None, []

    center = dict(center_row)
    center_days = set(parse_list(center.get('day', '')))
    center_times = parse_time_slots(parse_list(center.get('time', '')))
    total_day_slots = len(center_days)
    total_time_slots = len(center_times)

    instructor_rows = conn.execute(
        "SELECT * FROM instructor_applications ORDER BY id ASC"
    ).fetchall()
    conn.close()

    instructors = [dict(r) for r in instructor_rows]
    results = []

    for inst in instructors:
        try:
            choices_raw = inst.get('choices', '')
            choices = json.loads(choices_raw) if choices_raw else []
        except Exception:
            choices = []

        if not choices:
            continue

        direct_choice = None
        direct_apply_rank = None
        inst_region = ''
        for ch in choices:
            if not inst_region and ch.get('region'):
                inst_region = ch.get('region', '')
            if ch.get('region') == center_region and ch.get('org') == center_org:
                try:
                    rank = int(ch.get('rank', 99))
                    if direct_apply_rank is None or rank < direct_apply_rank:
                        direct_apply_rank = rank
                        direct_choice = ch
                except ValueError:
                    pass

        if direct_choice is None:
            continue

        all_days = set(direct_choice.get('days', []))
        all_times_raw = list(direct_choice.get('times', []))
        all_times = parse_time_slots(all_times_raw)

        matched_days = center_days & all_days
        missing_days = center_days - all_days
        day_coverage = len(matched_days) / total_day_slots if total_day_slots > 0 else 0
        matched_times = center_times & all_times
        time_coverage = len(matched_times) / total_time_slots if total_time_slots > 0 else 0

        score = 0
        score += day_coverage * 50
        score += time_coverage * 50

        rank_bonus_map = {1: 3.0, 2: 2.5, 3: 2.0, 4: 1.5, 5: 1.0}
        if direct_apply_rank is not None:
            score += rank_bonus_map.get(direct_apply_rank, 0)

        inst_name = str(inst.get('name', '')).strip()
        bonus_dict = bonus_data.get(inst_name, {})
        bonus_score, bonus_labels = calc_bonus_score(bonus_dict)
        interview_score = bonus_dict.get('interview_score', 0.0)
        score += bonus_score

        final_score = (score + interview_score) / 2

        reasons = []
        reasons.append(f'요일 충족률 {int(day_coverage*100)}% ({len(matched_days)}/{total_day_slots}일)')
        reasons.append(f'교시 충족률 {int(time_coverage*100)}% ({len(matched_times)}/{total_time_slots}교시)')
        if interview_score > 0:
            reasons.append(f'면접 최종점수 {round(interview_score, 1)}점 반영')
        if direct_apply_rank is not None:
            reasons.append(f'해당 센터 {direct_apply_rank}지망 직접 지원')
        if bonus_labels:
            for bl in bonus_labels:
                reasons.append(f'면접우대 가산: {bl}')
        if 0 < len(missing_days) <= 1:
            reasons.append(f'협의 시 {"·".join(sorted(missing_days))}요일 추가 가능성 있음')

        missing_days_str = '·'.join(sorted(missing_days)) if missing_days else ''
        matched_days_str = '·'.join(sorted(matched_days)) if matched_days else ''

        if len(missing_days) == 0 and time_coverage >= 1.0:
            script = f'"{matched_days_str} 전 교시 가능하신 것 확인했습니다. 바로 배정 가능하신가요?"'
        elif len(missing_days) == 0 and time_coverage < 1.0:
            missing_t_str = '·'.join(sorted(center_times - all_times))
            script = f'"{matched_days_str} 지원해주셨는데, {missing_t_str}도 가능하신가요?"'
        elif len(missing_days) == 1:
            script = f'"{matched_days_str} 지원해주셨는데, {missing_days_str}요일도 가능하신가요?"'
        elif len(missing_days) >= 2 and len(matched_days) > 0:
            script = (f'"{matched_days_str}·{missing_days_str} 중 저희가 현재 '
                      f'{missing_days_str}요일 자리만 남았습니다. '
                      f'{missing_days_str}요일만 강의하시는 것도 가능할까요?"')
        else:
            script = '"강의 가능 요일·시간 협의 필요"'

        results.append({
            'name': inst_name,
            'tel': inst.get('tel', ''),
            'role': inst.get('role', ''),
            'region': inst_region,
            'matched_days': sorted(matched_days),
            'matched_times': sorted(matched_times),
            'missing_days': sorted(missing_days),
            'direct_apply': direct_apply_rank is not None,
            'direct_apply_rank': direct_apply_rank,
            'bonus_score': bonus_score,
            'bonus_labels': bonus_labels,
            'interview_score': round(interview_score, 1),
            'raw_score': round(score, 1),
            'score': round(final_score, 1),
            'reasons': reasons,
            'script': script
        })

    results.sort(key=lambda x: x['score'], reverse=True)
    return center, {
        'days': sorted(center_days),
        'times': sorted(center_times)
    }, results[:20]
