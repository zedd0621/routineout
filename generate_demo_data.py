"""aiedu 데모 테넌트용 가짜 데이터 생성 스크립트.
실제 개인정보를 전혀 쓰지 않고, 원본 엑셀/DB와 동일한 구조로 더미 데이터를 새로 만든다.
로컬/서버 양쪽에서 한 번씩 실행해서 tenant_data/aiedu/ 아래에 데이터를 채운다.
"""
import json
import os
import random
from datetime import datetime, timedelta

from openpyxl import Workbook

from tenant_db import init_db, get_db, paycheck_folder, tenant_dir
import recommendation_engine as rec
import payroll_engine as pay

TENANT = 'aiedu'
INSTRUCTOR_COUNT = 100  # 기존 10명 → 10배 확장

SURNAMES = ['김', '이', '박', '최', '정', '강', '조', '윤', '한', '오',
            '서', '신', '권', '황', '안', '송', '류', '전', '홍', '문']
GIVEN_1 = ['민', '서', '도', '하', '지', '수', '은', '윤', '재', '현',
           '준', '아', '우', '연', '영', '나', '유', '태', '소', '진']
GIVEN_2 = ['준', '연', '윤', '은', '호', '아', '우', '민', '진', '율',
           '현', '빈', '서', '결', '결', '온', '결', '결', '결', '결']


def _gen_fake_instructors(n):
    rng = random.Random(7)  # 이름 생성은 별도 시드로 고정 — 매 실행 동일 결과 보장
    used_names = set()
    used_phones = set()
    instructors = []
    while len(instructors) < n:
        name = rng.choice(SURNAMES) + rng.choice(GIVEN_1) + rng.choice(GIVEN_2)
        if name in used_names:
            continue
        used_names.add(name)

        year = rng.randint(1965, 2001)
        month = rng.randint(1, 12)
        day = rng.randint(1, 28)
        birth = f'{year:04d}-{month:02d}-{day:02d}'

        while True:
            phone = f"010-{rng.randint(1000,9999)}-{rng.randint(1000,9999)}"
            if phone not in used_phones:
                used_phones.add(phone)
                break

        role = '주강사' if rng.random() < 0.6 else '보조강사'
        instructors.append({'name': name, 'birth': birth, 'tel': phone, 'role': role})
    return instructors


FAKE_INSTRUCTORS = _gen_fake_instructors(INSTRUCTOR_COUNT)

ORG_TYPES = ['복지관', '주민센터', '문화센터', '평생학습관', '청소년수련관', '노인종합복지관', '보건소', '도서관']
ORG_PREFIXES = ['중앙', '동부', '서부', '남부', '신도시']
ALL_COURSE_POOL = (rec.COURSES_BASIC + rec.COURSES_LIFE + rec.COURSES_DEEP)


def _gen_fake_centers():
    # 지역당 5개씩 만든다 — 드롭다운이 실제로 "선택할 게 있는" 수준이 되려면
    # 최소 이 정도는 있어야 한다는 피드백 반영.
    rng = random.Random(13)
    centers = []
    for region in rec.REGIONS:
        types = rng.sample(ORG_TYPES, k=5)
        for prefix, otype in zip(ORG_PREFIXES, types):
            org = f'가짜{region}{prefix}{otype}'
            days = ','.join(sorted(rng.sample(rec.ALL_DAYS, k=rng.randint(2, 4)),
                                    key=rec.ALL_DAYS.index))
            times = ','.join(sorted(rng.sample(rec.ALL_TIMES, k=rng.randint(1, 3)),
                                     key=rec.ALL_TIMES.index))
            months = ','.join(sorted(rng.sample(rec.ALL_MONTHS, k=rng.randint(1, 3)),
                                      key=rec.ALL_MONTHS.index))
            courses = ', '.join(rng.sample(ALL_COURSE_POOL, k=rng.randint(2, 4)))
            centers.append({
                'region': region,
                'org': org,
                'addr': f'경기도 {region}시 가짜로 {rng.randint(1, 200)}',
                'manager': rng.choice(SURNAMES) + '담당',
                'tel': f'031-{rng.randint(200,999)}-{rng.randint(1000,9999)}',
                'day': days,
                'time': times,
                'period': months,
                'course': courses,
                'room_count': str(rng.randint(1, 3)),
                'room_info': f'수용인원: {rng.choice([15,20,25,30])}명',
                'device': rng.choice(['데스크탑', '노트북', '태블릿', '데스크탑, 노트북']),
                'target': rng.choice(['어르신(60대 이상)', '중장년층(40~50대)', '일반성인', '장애인']),
                'headcount': f'{rng.randint(10,25)}명',
            })
    return centers


FAKE_CENTERS = _gen_fake_centers()

TIME_SLOTS = [
    (10, 0, 12, 0, '1교시(10:00~12:00)'),
    (13, 0, 15, 0, '2교시(13:00~15:00)'),
    (16, 0, 18, 0, '3교시(16:00~18:00)'),
]

WORKLOG_HEADER = [
    '사원번호', '이름', '지점', '직무', '날짜', '요일', '근무유형', '근무일정 템플릿',
    '근무일정\n시작시간', '근무일정\n종료시간', '출근시간', '(시급계산단위 적용)\n출근시간',
    '퇴근시간', '(시급계산단위 적용)\n퇴근시간', '출퇴근기록\n휴게시간', '(계획)\n자동추가된\n휴게시간',
    '(실제)\n(시급계산단위 적용)\n자동추가된\n휴게시간', '(계획)\n총 휴게시간', '(실제)\n총 휴게시간',
    '(실제)\n(시급계산단위 적용)\n총 휴게시간', '(계획)\n총 시간', '(실제)\n총 시간',
    '(실제)\n(시급계산단위 적용)\n총 시간', '(실제)\n강의시간', '주강사\n교육인원', '보조강사\n교육인원',
]


def make_worklog(target_month, out_path):
    year, month = map(int, target_month.split('-'))
    wb = Workbook()
    ws = wb.active
    ws.title = '실급여정산(근무일정 및 출퇴근기록 기반)'

    ws.cell(row=1, column=1, value='routineout 데모(가짜데이터)')
    ws.cell(row=1, column=3, value='야간근로 시작시간')
    ws.cell(row=1, column=4, value='22:00')
    ws.cell(row=2, column=3, value='야간근로 종료시간')
    ws.cell(row=2, column=4, value='06:00')
    for col, label in enumerate(WORKLOG_HEADER, start=1):
        ws.cell(row=3, column=col, value=label)

    weekday_kr = ['월', '화', '수', '목', '금', '토', '일']
    row_idx = 5

    import calendar
    days_in_month = calendar.monthrange(year, month)[1]
    all_dates = [datetime(year, month, d) for d in range(1, days_in_month + 1)]

    # 주(월요일 기준) 단위로 날짜를 묶는다 — 프로필별로 특정 주에 몰아서 배정하기 위함
    weeks = {}
    for dt in all_dates:
        if dt.weekday() == 6:  # 일요일 제외
            continue
        monday = dt - timedelta(days=dt.weekday())
        weeks.setdefault(monday, []).append(dt)
    # 월경계에 걸친 주(월요일이 이번 달이 아닌 주)는 build_payroll_summary의 중복지급
    # 방지 로직(같은 entries를 prior_entries로 재사용)에 걸려 주휴수당이 0으로 지워진다.
    # 데모에서 주휴수당이 확실히 보이도록, 월 안에 완전히 포함된 주만 후보로 쓴다.
    week_list = [weeks[k] for k in sorted(weeks.keys()) if k.month == month]

    # 강사별 근무 프로필: heavy(2명)=4대보험(월 60h+) 시연, medium(4명)=주휴수당(주 15h+) 시연,
    # light(4명)=둘 다 미달 시연 — 세 케이스가 전부 데모에서 보이도록 의도적으로 분포시킴
    profiles = ['heavy', 'heavy', 'medium', 'medium', 'medium', 'medium',
                'light', 'light', 'light', 'light']

    for idx, inst in enumerate(FAKE_INSTRUCTORS):
        profile = profiles[idx % len(profiles)]

        if profile == 'heavy':
            active_weeks = week_list[:min(4, len(week_list))]
            days_per_week = 6
            sessions_per_day = 2
        elif profile == 'medium':
            active_weeks = week_list[:1]
            days_per_week = 6
            sessions_per_day = 2
        else:
            active_weeks = week_list[:min(2, len(week_list))]
            days_per_week = 2
            sessions_per_day = 1

        for wk in active_weeks:
            wk_days = [d for d in wk if d.weekday() < 6]
            chosen_days = wk_days[:days_per_week]
            for date_obj in chosen_days:
                for s in range(sessions_per_day):
                    slot = TIME_SLOTS[s % len(TIME_SLOTS)]
                    sh, sm, eh, em, _ = slot
                    sched_start = date_obj.replace(hour=sh, minute=sm)
                    sched_end = date_obj.replace(hour=eh, minute=em)

                    late_in = random.random() < 0.15
                    early_out = random.random() < 0.15
                    actual_in = sched_start - timedelta(minutes=35) if late_in else sched_start
                    actual_out = sched_end + timedelta(minutes=35) if early_out else sched_end

                    sched_hours = (sched_end - sched_start).total_seconds() / 3600
                    headcount_main = random.randint(12, 28)
                    headcount_sub = random.randint(12, 28)

                    row = [None] * 26
                    row[0] = f'A{idx + 1:03d}'
                    row[1] = inst['name']
                    row[2] = FAKE_CENTERS[idx % len(FAKE_CENTERS)]['org']
                    row[3] = inst['role']
                    row[4] = date_obj
                    row[5] = weekday_kr[date_obj.weekday()]
                    row[6] = '정상근무'
                    row[7] = '기본템플릿'
                    row[8] = sched_start
                    row[9] = sched_end
                    row[10] = actual_in
                    row[11] = actual_in
                    row[12] = actual_out
                    row[13] = actual_out
                    row[14] = timedelta(0)
                    row[15] = timedelta(0)
                    row[16] = timedelta(0)
                    row[17] = timedelta(0)
                    row[18] = timedelta(0)
                    row[19] = timedelta(0)
                    row[20] = timedelta(hours=sched_hours)
                    row[21] = timedelta(hours=sched_hours)
                    row[22] = timedelta(hours=sched_hours)
                    row[23] = timedelta(hours=sched_hours)
                    row[24] = headcount_main
                    row[25] = headcount_sub

                    for col, val in enumerate(row, start=1):
                        ws.cell(row=row_idx, column=col, value=val)
                    row_idx += 1

    wb.save(out_path)


def make_instructor_list(out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.cell(row=1, column=1, value='성명(생년월일)')
    ws.cell(row=1, column=2, value='이름')
    ws.cell(row=1, column=3, value='연락처')
    for i, inst in enumerate(FAKE_INSTRUCTORS, start=2):
        ws.cell(row=i, column=1, value=f"{inst['name']} ({inst['birth']})")
        ws.cell(row=i, column=2, value=inst['name'])
        ws.cell(row=i, column=3, value=inst['tel'])
    wb.save(out_path)


def make_interview_results(out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = '면접결과 관리표'
    # 실제 파일의 헤더 영역(1~6행)은 병합/장식 위주라 데모에서는 최소 라벨만 채운다.
    ws.cell(row=6, column=15, value='최종합불 여부')
    ws.cell(row=6, column=18, value='기타 우대사항')
    ws.cell(row=6, column=21, value='보정값적용 최종점수')

    bonus_pool = ['청년', '신규강사', '취약계층', '청년/신규강사', '']
    row_idx = 7
    for inst in FAKE_INSTRUCTORS:
        passed = random.random() < 0.8
        ws.cell(row=row_idx, column=15, value=random.choice(bonus_pool))  # O(14, 0-idx)
        ws.cell(row=row_idx, column=16, value='합격' if passed else '불합격')  # P(15)
        ws.cell(row=row_idx, column=18, value=inst['name'])  # R(17)
        ws.cell(row=row_idx, column=21, value=round(random.uniform(70, 98), 1))  # U(20)
        row_idx += 1
    wb.save(out_path)


def seed_db():
    init_db(TENANT)
    conn = get_db(TENANT)

    conn.execute('DELETE FROM responses')
    conn.execute('DELETE FROM instructor_applications')
    for center in FAKE_CENTERS:
        conn.execute('''
            INSERT INTO responses (
                submitted_at, region, org, addr, manager, tel,
                room_count, room_info, device, target, headcount, course,
                recruit, recruit_channel, period, day, time,
                edubus, edubus_addr, edubus_date, etc, confirmed
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)
        ''', (
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            center['region'], center['org'], center['addr'], center['manager'], center['tel'],
            center['room_count'], center['room_info'], center['device'], center['target'],
            center['headcount'], center['course'],
            '자체모집', '현수막', center['period'], center['day'], center['time'],
            'N', '', '', '데모용 가짜 데이터'
        ))

    app_rng = random.Random(11)  # 지원 내역(요일·교시 부분지원)도 별도 시드로 고정
    for i, inst in enumerate(FAKE_INSTRUCTORS):
        center = FAKE_CENTERS[i % len(FAKE_CENTERS)]
        center_days = center['day'].split(',')
        center_times = center['time'].split(',')
        # 전원이 센터 요구조건을 100% 만족하면 추천 점수가 전부 동점이 되어 데모 효과가
        # 떨어진다 — 요일·교시 일부만 지원하는 인원을 섞어 점수 편차를 자연스럽게 만든다.
        days = center_days if app_rng.random() < 0.5 else app_rng.sample(
            center_days, k=app_rng.randint(1, len(center_days)))
        times = center_times if app_rng.random() < 0.5 else app_rng.sample(
            center_times, k=app_rng.randint(1, len(center_times)))
        choices = [{
            'rank': 1,
            'region': center['region'],
            'org': center['org'],
            'days': days,
            'times': times,
        }]
        conn.execute('''
            INSERT INTO instructor_applications (submitted_at, name, tel, role, choices)
            VALUES (?, ?, ?, ?, ?)
        ''', (
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            inst['name'], inst['tel'], inst['role'],
            json.dumps(choices, ensure_ascii=False)
        ))

    conn.commit()
    conn.close()


BANKS = ['국민은행', '신한은행', '우리은행', '하나은행', '농협은행', '카카오뱅크', '토스뱅크']


def seed_consents(target_month):
    """가상 데이터이므로 근무기록이 있는 전원이 급여내역을 확인·동의했다고 가정하고 채운다."""
    pf = paycheck_folder(TENANT)
    conn = get_db(TENANT)
    conn.execute('DELETE FROM paycheck_consents')

    worklog = pay.load_worklog(pf, target_month)
    prior = pay.load_worklog(pf, pay.prev_month_str(target_month))
    rng = random.Random(17)

    inst_by_name = {i['name']: i for i in FAKE_INSTRUCTORS}

    for name, entries in worklog.items():
        inst = inst_by_name.get(name)
        if not inst:
            continue
        daily_total = sum(e['daily_pay'] for e in entries)
        _, weekly_total = pay.calc_weekly_holiday(entries, target_month, prior_entries=prior.get(name, []))
        total_pay = daily_total + weekly_total
        is_four, _, _ = pay.judge_insurance(entries, target_month)
        insurance_type = '4대보험' if is_four else '2대보험(고용·산재)'
        tel_last4 = inst['tel'][-4:]
        bank_name = rng.choice(BANKS)
        account_number = f'{rng.randint(100,999)}-{rng.randint(10,99)}-{rng.randint(100000,999999)}'
        # consented_at: 정산월 말일 근처로 그럴듯하게
        consented_at = f'{target_month}-27 {rng.randint(9,18):02d}:{rng.randint(0,59):02d}:00'

        conn.execute('''
            INSERT INTO paycheck_consents
            (consented_at, name, birth, tel_last4, target_month, total_pay,
             ip_address, snapshot_json, bank_name, account_number, insurance_type)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            consented_at, name, inst['birth'], tel_last4, target_month, int(total_pay),
            f'{rng.randint(10,220)}.{rng.randint(0,255)}.{rng.randint(0,255)}.{rng.randint(0,255)}',
            json.dumps({'demo': True}, ensure_ascii=False),
            bank_name, account_number, insurance_type
        ))

    conn.commit()
    conn.close()


def main():
    random.seed(42)
    tdir = tenant_dir(TENANT)
    pf = paycheck_folder(TENANT)

    make_instructor_list(os.path.join(pf, 'instructor_list.xlsx'))
    make_worklog('2026-06', os.path.join(pf, 'worklog_2026-06.xlsx'))
    make_worklog('2026-07', os.path.join(pf, 'worklog_2026-07.xlsx'))
    make_interview_results(os.path.join(tdir, 'interview_results.xlsx'))
    seed_db()
    seed_consents('2026-07')
    print('demo data generated for tenant:', TENANT)


if __name__ == '__main__':
    main()
